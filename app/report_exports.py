"""Bounded, in-memory XLSX exports for the authorized EHF metrics projection."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any, Callable, Protocol

from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.citation_plots import citation_plot_points
from app.internal_preview import PreviewApplicantMetric


METRIC_HEADERS = (
    "Applicant",
    "Degree",
    "Age",
    "Academic age (years)",
    "Gender",
    "First-author papers",
    "Last-author papers",
    "Total papers",
    "h-index",
    "Total citations",
    "ORCID",
    "Google Scholar citations",
    "GS identity certainty",
)

_METRIC_ATTRIBUTES = (
    "applicant",
    "degree",
    "age",
    "academic_age",
    "gender",
    "first_author_papers",
    "last_author_papers",
    "total_papers",
    "h_index",
    "total_citations",
    "orcid",
    "google_scholar_citations",
    "identity_certainty",
)


@dataclass(frozen=True, slots=True)
class ReportExportMetadata:
    actor_identity: str
    actor_group: str
    generated_at_utc: datetime
    call_code: str = "EHF-2026"
    filters: str = "None"


class ReportAuditRepository(Protocol):
    def record(
        self,
        metadata: ReportExportMetadata,
        row_count: int,
        outcome: str,
        failure_stage: str | None = None,
    ) -> None: ...


class EmptyReportAuditRepository:
    def record(
        self,
        metadata: ReportExportMetadata,
        row_count: int,
        outcome: str,
        failure_stage: str | None = None,
    ) -> None:
        del metadata, row_count, outcome, failure_stage


class SqlReportAuditRepository:
    """Append export audit facts through the bounded SQL procedure only."""

    def __init__(self, connection_factory: Callable[[], Any]) -> None:
        self._connection_factory = connection_factory

    def record(
        self,
        metadata: ReportExportMetadata,
        row_count: int,
        outcome: str,
        failure_stage: str | None = None,
    ) -> None:
        connection = self._connection_factory()
        execute = getattr(connection, "execute", None)
        if execute is None:
            with connection as opened_connection:
                self._record_with_connection(
                    opened_connection, metadata, row_count, outcome, failure_stage
                )
            return
        self._record_with_connection(connection, metadata, row_count, outcome, failure_stage)

    @staticmethod
    def _record_with_connection(
        connection: Any,
        metadata: ReportExportMetadata,
        row_count: int,
        outcome: str,
        failure_stage: str | None,
    ) -> None:
        connection.execute(
            "EXEC dbo.RecordReportExportAudit @ActorIdentity=?, @ActorGroup=?, "
            "@RowCount=?, @Outcome=?, @FailureStage=?",
            metadata.actor_identity,
            metadata.actor_group,
            row_count,
            outcome,
            failure_stage,
        )
        connection.commit()


def safe_excel_text(value: str) -> str:
    """Keep untrusted strings from being interpreted as spreadsheet formulas."""
    return "'" + value if value.startswith(("=", "+", "-", "@")) else value


def build_metrics_workbook(
    records: tuple[PreviewApplicantMetric, ...], metadata: ReportExportMetadata
) -> bytes:
    """Build the approved metrics workbook entirely in memory."""
    workbook = Workbook()
    metrics = workbook.active
    metrics.title = "Applicant metrics"
    charts = workbook.create_sheet("Charts")
    export_metadata = workbook.create_sheet("Export metadata")

    _write_metrics_sheet(metrics, records)
    _write_charts_sheet(charts, records)
    _write_metadata_sheet(export_metadata, metadata, len(records))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_metrics_sheet(sheet, records: tuple[PreviewApplicantMetric, ...]) -> None:  # type: ignore[no-untyped-def]
    sheet["A1"] = "EHF 2026 applicant metrics"
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color="17365D")
    sheet["A2"] = "Prepared from the current authorized application-metrics projection."
    sheet["A4"] = "Notes"
    sheet["A4"].font = Font(name="Aptos", bold=True)
    sheet["A5"] = "NR means not recorded in the source register."
    sheet["A6"] = (
        "Academic age is the recorded career duration; citation plots use total citations "
        "with Google Scholar citations as the fallback."
    )

    header_row = 8
    for column, header in enumerate(METRIC_HEADERS, start=1):
        cell = sheet.cell(header_row, column, header)
        cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row_number, record in enumerate(records, start=header_row + 1):
        for column, attribute in enumerate(_METRIC_ATTRIBUTES, start=1):
            raw = getattr(record, attribute)
            value = "NR" if raw in (None, "") else raw
            if isinstance(value, str):
                value = safe_excel_text(value)
            cell = sheet.cell(row_number, column, value)
            cell.font = Font(name="Aptos", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    last_row = max(header_row + 1, header_row + len(records))
    if not records:
        for column in range(1, len(METRIC_HEADERS) + 1):
            sheet.cell(last_row, column, None)
    table = Table(displayName="EHFApplicantMetrics", ref=f"A{header_row}:M{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    sheet.add_table(table)
    sheet.freeze_panes = f"A{header_row + 1}"
    widths = (28, 14, 10, 18, 12, 16, 16, 14, 10, 16, 23, 22, 22)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    sheet.sheet_view.showGridLines = False


def _write_charts_sheet(sheet, records: tuple[PreviewApplicantMetric, ...]) -> None:  # type: ignore[no-untyped-def]
    sheet["A1"] = "EHF 2026 citation charts"
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color="17365D")
    chart_headers = ("Applicant", "Anagraphic age", "Academic age", "Citations")
    for column, header in enumerate(chart_headers, start=1):
        cell = sheet.cell(3, column, header)
        cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row_number, record in enumerate(records, start=4):
        citations = (
            record.total_citations
            if record.total_citations is not None
            else record.google_scholar_citations
        )
        values = (record.applicant, record.age, record.academic_age, citations)
        for column, value in enumerate(values, start=1):
            if isinstance(value, str):
                value = safe_excel_text(value)
            sheet.cell(row_number, column, value)

    for title, x_column, age_field, anchor in (
        ("Citations by anagraphic age", 2, "age", "F3"),
        ("Citations by academic age", 3, "academic_age", "F20"),
    ):
        chart = ScatterChart()
        chart.title = title
        chart.style = 13
        chart.height = 8
        chart.width = 15
        chart.legend = None
        chart.varyColors = False
        chart.x_axis.title = "Age (years)"
        chart.y_axis.title = "Total citations"
        points = citation_plot_points(records, age_field)
        if points:
            ages = tuple(point.age for point in points)
            citations = tuple(point.citations for point in points)
            age_span = max(max(ages) - min(ages), 1.0)
            chart.x_axis.scaling.min = max(0.0, min(ages) - age_span * 0.04)
            chart.x_axis.scaling.max = max(ages) + age_span * 0.32
            chart.y_axis.scaling.min = 0.0
            chart.y_axis.scaling.max = max(max(citations) * 1.12, 1.0)
        for point in points:
            source_row = 4 + point.source_index
            x_values = Reference(
                sheet, min_col=x_column, min_row=source_row, max_row=source_row
            )
            y_values = Reference(
                sheet, min_col=4, min_row=source_row, max_row=source_row
            )
            series = Series(y_values, x_values, title=point.surname)
            series.marker.symbol = "circle"
            series.marker.size = 7
            series.marker.graphicalProperties.solidFill = point.color[1:]
            series.marker.graphicalProperties.line.solidFill = "1D2525"
            series.graphicalProperties.line.noFill = True
            if point.labelled:
                series.dLbls = DataLabelList(
                    dLblPos="r",
                    showLegendKey=False,
                    showVal=False,
                    showCatName=False,
                    showSerName=True,
                )
            chart.series.append(series)
        sheet.add_chart(chart, anchor)

    sheet.column_dimensions["A"].width = 28
    for column in ("B", "C", "D"):
        sheet.column_dimensions[column].width = 18
    sheet.sheet_view.showGridLines = False


def _write_metadata_sheet(
    sheet, metadata: ReportExportMetadata, row_count: int  # type: ignore[no-untyped-def]
) -> None:
    sheet["A1"] = "Export metadata"
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color="17365D")
    values = (
        ("Call", metadata.call_code),
        ("Exporting identity", metadata.actor_identity),
        ("Authorization group", metadata.actor_group),
        ("Generated at (UTC)", metadata.generated_at_utc.isoformat()),
        ("Row count", row_count),
        ("Filters", metadata.filters),
        ("Notice", "Confidential EHF working material for authorized internal use only."),
    )
    for row_number, (label, value) in enumerate(values, start=3):
        sheet.cell(row_number, 1, label).font = Font(name="Aptos", bold=True)
        sheet.cell(row_number, 2, safe_excel_text(value) if isinstance(value, str) else value)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 70
    sheet.sheet_view.showGridLines = False
