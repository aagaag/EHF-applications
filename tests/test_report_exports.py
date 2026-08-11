"""Contracts for the authorized EHF metrics workbook and download route."""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.config import Settings
from app import report_exports
from app.identity import AuthenticatedIdentity
from app.internal_preview import PreviewApplicantMetric
from app.main import ReadinessChecks, create_app
from app.navigation import INTERNAL_GROUPS
from app.preferences import Identity
from app.report_exports import METRIC_HEADERS, ReportExportMetadata, build_metrics_workbook


def _records() -> tuple[PreviewApplicantMetric, ...]:
    return (
        PreviewApplicantMetric(
            applicant="=DANGEROUS()",
            degree="PhD",
            age=36,
            academic_age=8.5,
            gender=None,
            first_author_papers=7,
            last_author_papers=2,
            total_papers=18,
            h_index=12,
            total_citations=640,
            orcid="0000-0002-1825-0097",
            google_scholar_citations=710,
            identity_certainty="High",
        ),
        PreviewApplicantMetric(
            applicant="Zoë Example",
            degree=None,
            age=31,
            academic_age=4,
            total_papers=9,
            google_scholar_citations=125,
        ),
    )


def _metadata(group: str = INTERNAL_GROUPS.trustees) -> ReportExportMetadata:
    return ReportExportMetadata(
        actor_identity="cloudflare:stable-subject",
        actor_group=group,
        generated_at_utc=datetime(2026, 8, 10, 11, 30, tzinfo=UTC),
    )


def _contrast_ratio(first: str, second: str) -> float:
    def luminance(value: str) -> float:
        channels = [int(value[index:index + 2], 16) / 255 for index in (0, 2, 4)]
        linear = [
            channel / 12.92
            if channel <= 0.04045
            else ((channel + 0.055) / 1.055) ** 2.4
            for channel in channels
        ]
        return 0.2126 * linear[0] + 0.7152 * linear[1] + 0.0722 * linear[2]

    lighter, darker = sorted((luminance(first), luminance(second)), reverse=True)
    return (lighter + 0.05) / (darker + 0.05)


def test_workbook_preserves_the_approved_metrics_contract_and_native_charts() -> None:
    workbook = load_workbook(BytesIO(build_metrics_workbook(_records(), _metadata())))

    assert workbook.sheetnames == ["Applicant metrics", "Charts", "Export metadata"]
    metrics = workbook["Applicant metrics"]
    header_row = next(
        row for row in metrics.iter_rows() if tuple(cell.value for cell in row) == METRIC_HEADERS
    )
    assert tuple(cell.value for cell in header_row) == METRIC_HEADERS
    assert metrics.cell(header_row[0].row + 1, 1).value == "'=DANGEROUS()"
    assert metrics.cell(header_row[0].row + 1, 3).value == 36
    assert isinstance(metrics.cell(header_row[0].row + 1, 4).value, float)
    assert metrics.cell(header_row[0].row + 1, 5).value == "NR"
    assert metrics.cell(header_row[0].row + 2, 1).value == "Zoë Example"
    assert metrics.tables
    assert metrics.auto_filter.ref is None
    assert len(workbook["Charts"]._charts) == 2
    assert all(chart.legend is None for chart in workbook["Charts"]._charts)

    all_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    ).casefold()
    for forbidden in ("email", "token", "password", "recommendation", "document path"):
        assert forbidden not in all_text


def test_workbook_charts_use_distinct_shared_colors_and_top_15_surname_labels() -> None:
    """Break caught: exported plots could collapse applicants into one unlabeled series."""
    records = tuple(
        PreviewApplicantMetric(
            applicant=f"Given Surname{index:02d}",
            age=30 + index,
            academic_age=3 + index,
            total_citations=index,
        )
        for index in range(18)
    )
    workbook = load_workbook(BytesIO(build_metrics_workbook(records, _metadata())))
    charts = workbook["Charts"]._charts

    assert len(charts) == 2
    colors_by_chart: list[list[str]] = []
    for chart, ages in zip(
        charts,
        (
            [float(record.age) for record in records if record.age is not None],
            [
                float(record.academic_age)
                for record in records
                if record.academic_age is not None
            ],
        ),
        strict=True,
    ):
        assert len(chart.series) == 18
        colors = [
            series.marker.graphicalProperties.solidFill.srgbClr
            for series in chart.series
        ]
        colors_by_chart.append(colors)
        assert len(set(colors)) == 18
        assert all(
            _contrast_ratio(
                series.marker.graphicalProperties.line.solidFill.srgbClr,
                "FFFFFF",
            ) >= 3
            for series in chart.series
        )
        assert all(series.graphicalProperties.line.noFill for series in chart.series)
        assert all(
            series.xVal.numRef.f.endswith(f"${4 + index}")
            for index, series in enumerate(chart.series)
        )

        labelled = [series for series in chart.series if series.dLbls is not None]
        assert len(labelled) == 15
        assert {series.tx.v for series in labelled} == {
            f"Surname{index:02d}" for index in range(3, 18)
        }
        assert all(series.dLbls.showSerName for series in labelled)
        assert all(series.dLbls.dLblPos == "r" for series in labelled)
        assert chart.x_axis.scaling.max >= max(ages) + (max(ages) - min(ages)) * 0.2
        assert chart.y_axis.scaling.max > max(
            record.total_citations for record in records
        )

    assert colors_by_chart[1] == colors_by_chart[0]


def test_export_metadata_is_bounded_and_identifies_the_export() -> None:
    workbook = load_workbook(BytesIO(build_metrics_workbook(_records(), _metadata())))
    values = {
        row[0].value: row[1].value
        for row in workbook["Export metadata"].iter_rows(min_col=1, max_col=2)
        if row[0].value
    }

    assert values["Call"] == "EHF-2026"
    assert values["Exporting identity"] == "cloudflare:stable-subject"
    assert values["Authorization group"] == "EHF-Trustees"
    assert values["Row count"] == 2
    assert values["Filters"] == "None"


class RecordingMetricRepository:
    def __init__(self) -> None:
        self.requested_roles: list[str] = []

    def load(self, canonical_group: str) -> tuple[PreviewApplicantMetric, ...]:
        self.requested_roles.append(canonical_group)
        return _records()


class RecordingAuditRepository:
    def __init__(self) -> None:
        self.events: list[tuple[ReportExportMetadata, int, str, str | None]] = []

    def record(
        self,
        metadata: ReportExportMetadata,
        row_count: int,
        outcome: str,
        failure_stage: str | None = None,
    ) -> None:
        self.events.append((metadata, row_count, outcome, failure_stage))


def _client(
    group: str,
    repository: RecordingMetricRepository,
    audit_repository: object | None = None,
    *,
    raise_server_exceptions: bool = True,
) -> TestClient:
    principal = AuthenticatedIdentity(
        Identity("cloudflare:stable-subject", "person@example.org", "Example Person"),
        frozenset({group}),
    )
    return TestClient(
        create_app(
            Settings.from_environment({"EHF_ALLOWED_HOST": "localhost"}),
            readiness_checks=ReadinessChecks(lambda _: None, lambda _: None),
            identity_resolver=lambda _request: principal,
            metric_repository=repository,
            report_audit_repository=audit_repository,
        ),
        base_url="http://localhost",
        raise_server_exceptions=raise_server_exceptions,
    )


def test_authorized_download_uses_the_role_scoped_projection() -> None:
    repository = RecordingMetricRepository()
    audit = RecordingAuditRepository()
    response = _client(INTERNAL_GROUPS.administrators, repository, audit).get(
        "/internal/reports/metrics.xlsx"
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"] == 'attachment; filename="ehf-2026.xlsx"'
    assert "no-store" in response.headers["cache-control"]
    assert repository.requested_roles == [INTERNAL_GROUPS.administrators]
    assert [(event[2], event[3]) for event in audit.events] == [("COMPLETED", None)]
    workbook = load_workbook(BytesIO(response.content))
    assert workbook["Applicant metrics"].tables
    assert len(workbook["Charts"]._charts) == 2


def test_unauthenticated_download_fails_closed() -> None:
    app = create_app(
        Settings.from_environment({"EHF_ALLOWED_HOST": "localhost"}),
        readiness_checks=ReadinessChecks(lambda _: None, lambda _: None),
    )
    assert TestClient(app, base_url="http://localhost").get(
        "/internal/reports/metrics.xlsx"
    ).status_code == 404


def test_workbook_failure_is_audited_without_returning_bytes(monkeypatch) -> None:  # type: ignore[no-untyped-def]
    repository = RecordingMetricRepository()
    audit = RecordingAuditRepository()
    monkeypatch.setattr(
        "app.main.build_metrics_workbook",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError("sensitive failure")),
    )

    response = _client(
        INTERNAL_GROUPS.trustees,
        repository,
        audit,
        raise_server_exceptions=False,
    ).get("/internal/reports/metrics.xlsx")

    assert response.status_code == 500
    assert "sensitive failure" not in response.text
    assert response.headers["content-type"].startswith("application/json")
    assert [(event[1], event[2], event[3]) for event in audit.events] == [
        (len(_records()), "FAILED", "workbook-generation")
    ]


def test_sql_audit_repository_calls_only_the_bounded_procedure() -> None:
    calls: list[tuple[object, ...]] = []
    commits: list[bool] = []

    class Connection:
        def execute(self, *arguments: object) -> None:
            calls.append(arguments)

        def commit(self) -> None:
            commits.append(True)

    repository = report_exports.SqlReportAuditRepository(lambda: Connection())
    repository.record(_metadata(), 2, "COMPLETED")

    assert len(calls) == 1
    assert calls[0][0] == (
        "EXEC dbo.RecordReportExportAudit @ActorIdentity=?, @ActorGroup=?, "
        "@RowCount=?, @Outcome=?, @FailureStage=?"
    )
    assert calls[0][1:] == (
        "cloudflare:stable-subject", "EHF-Trustees", 2, "COMPLETED", None
    )
    assert commits == [True]
